import random
from random import shuffle
from openpyxl import Workbook

# creating workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "One Player"
ws.page_setup.fitToWidth = 1

# defining n for games and finish
fo = int(input("finish order: ")) #default 1 000 000
go = int(input("game order: ")) #default 1 000

# roll values are values from a six by six grid for all dice rolls
rv = [2,3,4,5,6,7,
      3,4,5,6,7,8,
      4,5,6,7,8,9,
      5,6,7,8,9,10,
      6,7,8,9,10,11,
      7,8,9,10,11,12]

# defining tiles
tiles = ['Start',
         'Granada',
         'Seville',
         'Madrid',
         'Bali',
         'Hong Kong',
         'Beijing',
         'Shanghai',
         'Lost Island',
         'Venice',
         'Milan',
         'Rome',
         'Chance',
         'Hamburg',
         'Cyprus',
         'Berlin',
         'World Championships',
         'London',
         'Dubai',
         'Sydney',
         'Chance',
         'Chicago',
         'Las Vegas',
         'New York',
         'World Tour',
         'Nice',
         'Lyon',
         'Paris',
         'Chance',
         'Sochi',
         'Tax',
         'Moscow']

# placing defined tiles in cells
for i in range(1, 33):
    _ = ws.cell(column=1, row=i, value=tiles[i-1])

def bt(finish_order, games_order, rollvalues):
     
    finish = finish_order
    games = games_order
    rv = rollvalues
     
    squares = []
     
    while len(squares) < 32:
        squares.append(0)
    
    games_finished = 0
    
    while games_finished < games:
        
        # deck of chance cards, everything below 33 is a go to card
        master_chance = [0,8,16,24,30,33,33,33,33,33,33,33,33,33,33,33]
        chance = [i for i in master_chance]
        shuffle(chance)
        
        doubles = 0
        doubles_ = 0
        
        position = 0
        
        gos = 0
        
        while gos < finish:
            #rolling dice
            diceroll = int(36*random.random())
             
            if diceroll in [0,7,14,21,28,35]:    # these are the dice index values for double rolls
                doubles += 1
            else:
                doubles = 0
            if doubles >= 3:
                position = 8    # making sure you can't trow more then 3 doubles in a row
                doubles_ += 1
            else:
                # getting new position from dice roll
                position = (position + rv[diceroll])%32

                # if landed on a chance tile, pick a card from the deck
                if position in [12,20,28]:  # chance tiles
                    chance_card = chance.pop(0)
                    if len(chance) == 0:
                        chance = [i for i in master_chance]     # reshuffeling when deck is empty
                        shuffle(chance)
                    if chance_card != 33:
                        position = chance_card

            # inserting position into the square array
            squares.insert(position,(squares.pop(position)+1))
             
            gos += 1
        
        games_finished += 1
        print("Games finished: ", games_finished)

        if games_finished == games:
            for i in range(1,33):
                _ = ws.cell(column=2, row=i, value=squares[i-1])

            ws['A33'] = "Total: "
            ws['B33'] = "=SUM(B1:B32)"
            
            ws['E1'] = doubles_

    return squares

bt(fo, go, rv)

wb.save('BusinessTour.xlsx')